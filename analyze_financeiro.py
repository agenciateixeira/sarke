#!/usr/bin/env python3
import pandas as pd
import os
from openpyxl import load_workbook

# Caminho do arquivo
file_path = os.path.expanduser("~/Downloads/ADM - Prime N.S. Final (1).xlsx")

# Carregar o workbook para ver as abas
wb = load_workbook(file_path, read_only=True, data_only=True)

print("=" * 60)
print("ANÁLISE DA PLANILHA FINANCEIRA - ADM PRIME")
print("=" * 60)

print("\n📋 ABAS DISPONÍVEIS NA PLANILHA:")
print("-" * 40)
for i, sheet_name in enumerate(wb.sheetnames, 1):
    print(f"{i:2}. {sheet_name}")

print("\n" + "=" * 60)
print("ANÁLISE DETALHADA DE CADA ABA:")
print("=" * 60)

# Analisar cada aba
for sheet_name in wb.sheetnames:
    print(f"\n\n{'=' * 60}")
    print(f"📊 ABA: {sheet_name}")
    print("=" * 60)

    try:
        # Ler a aba
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

        # Mostrar dimensões
        print(f"📏 Dimensões: {df.shape[0]} linhas x {df.shape[1]} colunas")

        # Procurar cabeçalho (linha com palavras-chave financeiras)
        header_row = -1
        financial_keywords = [
            'data', 'valor', 'descrição', 'descricao',
            'entrada', 'saída', 'saida', 'categoria',
            'fornecedor', 'cliente', 'pagamento', 'recebimento',
            'total', 'saldo', 'banco', 'conta'
        ]

        for idx, row in df.iterrows():
            if idx > 30:  # Limitar busca
                break
            row_text = ' '.join(str(cell).lower() for cell in row if pd.notna(cell))
            if any(keyword in row_text for keyword in financial_keywords):
                header_row = idx
                print(f"\n✅ Cabeçalho encontrado na linha {idx + 1}")

                # Mostrar colunas encontradas
                print("\n📑 COLUNAS IDENTIFICADAS:")
                cols = [str(cell) for cell in row if pd.notna(cell)]
                for i, col in enumerate(cols, 1):
                    if col != 'nan':
                        print(f"   {i:2}. {col}")
                break

        # Se encontrou cabeçalho, analisar dados
        if header_row >= 0:
            # Re-ler com cabeçalho
            df_data = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)

            print("\n📊 RESUMO DOS DADOS:")
            print(f"   • Total de registros: {len(df_data) - 1}")

            # Identificar colunas importantes
            value_cols = []
            date_cols = []
            text_cols = []

            for col in df_data.columns:
                if pd.api.types.is_numeric_dtype(df_data[col]):
                    if df_data[col].sum() != 0:
                        value_cols.append(col)
                elif 'data' in str(col).lower():
                    date_cols.append(col)
                else:
                    text_cols.append(col)

            if value_cols:
                print(f"\n💰 COLUNAS DE VALORES:")
                for col in value_cols:
                    total = df_data[col].sum()
                    if total != 0:
                        print(f"   • {col}: R$ {total:,.2f}")

            if date_cols:
                print(f"\n📅 COLUNAS DE DATA:")
                for col in date_cols:
                    print(f"   • {col}")

            # Mostrar primeiras linhas de dados
            print(f"\n📋 AMOSTRA DE DADOS (primeiras 3 linhas):")
            data_start = header_row + 1
            sample_rows = df.iloc[data_start:data_start + 3]
            for idx, row in sample_rows.iterrows():
                non_empty = [(i, val) for i, val in enumerate(row) if pd.notna(val) and str(val) != 'nan']
                if non_empty:
                    print(f"\n   Linha {idx + 1}:")
                    for col_idx, val in non_empty[:5]:  # Mostrar até 5 colunas
                        print(f"      Col {col_idx + 1}: {str(val)[:50]}...")

        else:
            print("⚠️  Cabeçalho não encontrado - pode ser aba de resumo ou configuração")

            # Tentar identificar o tipo de conteúdo
            non_empty_cells = []
            for idx, row in df.iterrows():
                if idx > 20:
                    break
                for cell in row:
                    if pd.notna(cell) and str(cell) != 'nan':
                        non_empty_cells.append(str(cell))

            if non_empty_cells:
                print("\n📝 CONTEÚDO ENCONTRADO (amostra):")
                for i, content in enumerate(non_empty_cells[:10], 1):
                    print(f"   {i}. {content[:60]}...")

    except Exception as e:
        print(f"❌ Erro ao processar aba: {e}")

print("\n\n" + "=" * 60)
print("RESUMO FINAL:")
print("=" * 60)

# Identificar a aba principal de dados financeiros
print("\n🎯 ANÁLISE DE ESTRUTURA:")
print("   • Esta planilha parece conter dados administrativos/financeiros")
print("   • Múltiplas abas com diferentes aspectos do controle")
print("   • Verificar se há abas específicas para:")
print("     - Fluxo de Caixa")
print("     - Contas a Pagar/Receber")
print("     - Centro de Custos")
print("     - Relatórios Consolidados")

wb.close()