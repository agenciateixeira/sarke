#!/usr/bin/env python3
import pandas as pd
import os
from openpyxl import load_workbook

# Caminho do arquivo
file_path = os.path.expanduser("~/Downloads/Cronograma - Viale Conveniencia.xlsx")

# Carregar o workbook para ver as abas
wb = load_workbook(file_path, read_only=True)
print("=" * 50)
print("ABAS DISPONÍVEIS NA PLANILHA:")
print("=" * 50)
for i, sheet_name in enumerate(wb.sheetnames, 1):
    print(f"{i}. {sheet_name}")

print("\n" + "=" * 50)
print("ANÁLISE DE CADA ABA:")
print("=" * 50)

# Analisar cada aba
for sheet_name in wb.sheetnames:
    print(f"\n📋 ABA: {sheet_name}")
    print("-" * 30)

    try:
        # Ler a aba
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

        # Mostrar dimensões
        print(f"Dimensões: {df.shape[0]} linhas x {df.shape[1]} colunas")

        # Procurar cabeçalho (linha com palavras-chave)
        header_row = -1
        keywords = ['item', 'descrição', 'descricao', 'local', 'quantidade', 'valor', 'data', 'cronograma']

        for idx, row in df.iterrows():
            if idx > 20:  # Limitar busca às primeiras 20 linhas
                break
            row_text = ' '.join(str(cell).lower() for cell in row if pd.notna(cell))
            if any(keyword in row_text for keyword in keywords):
                header_row = idx
                print(f"Cabeçalho encontrado na linha {idx + 1}")
                print(f"Colunas: {list(df.iloc[idx])}")
                break

        # Se encontrou cabeçalho, mostrar algumas linhas de dados
        if header_row >= 0 and header_row + 1 < len(df):
            print(f"\nPrimeiras 5 linhas de dados:")
            data_rows = df.iloc[header_row + 1:header_row + 6]
            for idx, row in data_rows.iterrows():
                # Filtrar apenas células não vazias
                non_empty = [(i, val) for i, val in enumerate(row) if pd.notna(val) and str(val).strip()]
                if non_empty:
                    print(f"  Linha {idx + 1}: {non_empty[:5]}")  # Mostrar apenas primeiras 5 colunas não vazias

        # Contar linhas com dados
        data_count = 0
        if header_row >= 0:
            for idx in range(header_row + 1, len(df)):
                if df.iloc[idx].notna().any():
                    data_count += 1
        print(f"\nTotal de linhas com dados: {data_count}")

    except Exception as e:
        print(f"Erro ao processar aba: {e}")

print("\n" + "=" * 50)
print("ANÁLISE ESPECÍFICA - VERIFICANDO MÚLTIPLOS ITENS POR DIA:")
print("=" * 50)

# Focar na primeira aba que parece ser o cronograma
try:
    first_sheet = wb.sheetnames[0]
    df = pd.read_excel(file_path, sheet_name=first_sheet, header=None)

    # Procurar padrões de data
    date_pattern_found = False
    for idx, row in df.iterrows():
        if idx > 100:  # Limitar análise
            break
        for cell in row:
            if pd.notna(cell):
                cell_str = str(cell)
                # Verificar se parece uma data
                if '/' in cell_str or '-' in cell_str:
                    if any(char.isdigit() for char in cell_str):
                        print(f"Possível data na linha {idx + 1}: {cell_str}")
                        date_pattern_found = True
                        break
        if date_pattern_found and idx < 50:
            date_pattern_found = False  # Continuar procurando

except Exception as e:
    print(f"Erro na análise específica: {e}")

wb.close()